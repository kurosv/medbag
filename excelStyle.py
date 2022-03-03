import openpyxl as xl
from openpyxl import utils
from openpyxl.styles import Font, Alignment, Border, Side


# ================================================
# Excelスタイル設定クラス
#  Target : 作成したExcelファイル・シート
# TODO ループ最適化：余裕あるとき
#  -> 各設定ごとループ -> ループ一括化
# ================================================
class ExcelStyle:
    def __init__(self, exfile, exsheet):
        # === CONST ===
        self.INDEX_COLUMN_WIDTH = 60  # 設定用カラム幅：Index列
        self.CHK_COLUMN_WIDTH = 3  # 設定用カラム幅：Index列
        self.ROW_IDX = "1"  # INDEX行位置
        self.NAME_COL_IDX = "A"  # 医薬品名列位置
        self.FIXED_ROW = "A2"
        # 3カ月連続処方の判定用
        self.COL_CHK = 2  # 埋め込み列
        self.COL_CHK_IDX = 'B'
        self.COL_CHK_IDX_NAME = "・"  # CHK INDEX名
        self.CHK_STR = "〇"
        # AVE埋め込み用
        self.COL_AVE = 3  # 平均埋め込み列
        self.COL_AVE_IDX = 'C'
        self.COL_AVE_RNG_INI = "D"  # 平均範囲指定：スタート位置
        self.COL_AVE_IDX_NAME = "平均"  # AVE INDEX名
        self.AVE_STR = "=AVERAGE("  # AVE本体

        # font設定
        self.font_index = Font(name="Arial", bold=True)
        self.font_def = Font(name="Arial")
        # ファイル/シート保持用
        self.file = exfile
        self.sheet = exsheet
        # 最大列数保持
        self.max_cidx = 0
        self.max_ridx = 0

    # ===============================================
    # スタイルセット
    # ===============================================
    def set_style(self):
        try:
            # read input xlsx
            wb = xl.load_workbook(filename=self.file)
            ws = wb[self.sheet]
            self.max_cidx = ws.max_column
            self.max_ridx = ws.max_row
        except FileNotFoundError as e:
            print(e)
            print(type(e))
        else:
            # チェック埋め込み
            self.set_chk(ws)
            # AVE埋め込み
            self.set_ave(ws)
            # font設定
            self.set_font(ws)
            # カラム幅設定
            self.set_columnwidth(ws)
            # 罫線設定
            self.set_border(ws)
            # フィルタ設定
            self.set_filter(ws)
            # 先頭行固定
            ws.freeze_panes = self.FIXED_ROW

            # save xlsx file
            wb.save(self.file)
        finally:
            pass

    # ===============================================
    # フォント設定：Arial
    # ===============================================
    def set_font(self, ws):
        # write in sheet
        for row in ws:
            for cell in row:
                if cell.col_idx == 1:
                    ws[cell.coordinate].font = self.font_index
                    # ついでに位置調整
                    cell.alignment = Alignment(horizontal='left',
                                               vertical='center',
                                               wrap_text=False)
                if cell.col_idx == 2:
                    ws[cell.coordinate].font = self.font_index
                    # ついでに位置調整
                    cell.alignment = Alignment(horizontal='center',
                                               vertical='center',
                                               wrap_text=False)

                else:
                    ws[cell.coordinate].font = self.font_def
                    cell.alignment = Alignment(horizontal='right',
                                               vertical='center',
                                               wrap_text=False)

    # ===============================================
    # カラム幅設定
    #  Index列 : 幅60に設定
    #  Check列 : 幅2に設定
    # ===============================================
    def set_columnwidth(self, ws):
        # set column width
        # for col in ws.columns:
        # for cell in col:
        ws.column_dimensions[self.NAME_COL_IDX].width = self.INDEX_COLUMN_WIDTH
        ws.column_dimensions[self.COL_CHK_IDX].width = self.CHK_COLUMN_WIDTH

    # ===============================================
    # 罫線設定
    # ===============================================
    def set_border(self, ws):
        # set side style
        side_index = Side(style='double', color='000000')
        side_def = Side(style='thin', color='000000')

        # set border
        border_index = Border(top=side_def, bottom=side_index, left=side_def, right=side_def)
        border_def = Border(bottom=side_def, left=side_def, right=side_def)

        # write in sheet
        for row in ws:
            for cell in row:
                # Index有効   範囲で設定
                if cell.col_idx <= self.max_cidx:
                    # Index行のみ別指定
                    if cell.row == 1:
                        ws[cell.coordinate].border = border_index
                    # 2行目以降
                    else:
                        ws[cell.coordinate].border = border_def

    # ===============================================
    # 平均行埋め込み設定
    # ===============================================
    def set_ave(self, ws):
        # 平均列の挿入
        ws.insert_cols(self.COL_AVE)
        # 列名設定
        ws[self.COL_AVE_IDX + self.ROW_IDX].value = self.COL_AVE_IDX_NAME
        # 最大列更新
        self.max_cidx = ws.max_column

        # AVE埋め込み
        for row in ws:
            idx_r = row[0]
            r = idx_r.row
            # 空白行は飛ばし
            if idx_r.value is None or not str(idx_r.value).strip():
                continue
            # ヘッダ飛ばし
            elif r < 2:
                continue

            c = ws.cell(row=r, column=self.max_cidx).column_letter

            # AVE作成
            # "=AVE(C0:X0)"
            c_start = self.COL_AVE_RNG_INI + str(r) + ":"
            c_end = c + str(r) + ")"
            ave_str = self.AVE_STR + c_start + c_end

            ws[self.COL_AVE_IDX + str(r)].value = ave_str

    # ===============================================
    # チェック行埋め込み設定
    # ===============================================
    def set_chk(self, ws):
        # チェック列の挿入
        ws.insert_cols(self.COL_CHK)
        # 列名設定
        ws[self.COL_CHK_IDX + self.ROW_IDX].value = self.COL_CHK_IDX_NAME
        # 最大列更新
        self.max_cidx = ws.max_column

        # CHK判定・埋め込み
        for row in ws:
            idx_r = row[0]
            r = idx_r.row
            # 空白行は飛ばし
            if idx_r.value is None or not str(idx_r.value).strip():
                continue
            # ヘッダ飛ばし
            elif r < 2:
                continue

            # 連続処方判定用フラグ
            f = True
            for cell in row:
                if cell.value == 0:
                    f = False
                    break

            if f:
                ws[self.COL_CHK_IDX + str(r)].value = self.CHK_STR
            else:
                continue

    # ===============================================
    # フィルタ設定
    # ===============================================
    def set_filter(self, ws):
        max_col_letter = utils.get_column_letter(self.max_cidx)
        filter_range = "A1:" + max_col_letter + str(self.max_ridx)
        ws.auto_filter.ref = filter_range


if __name__ == '__main__':
    file = 'C:\\Users\\iruka\\Desktop\\ph_aggre\\ph_aggre.xlsx'
    sname = '2021_07-08'
    es = ExcelStyle(file, sname)
    es.set_style()
