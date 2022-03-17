"""
makemb
  TODO 薬袋作成
  ・患者ごとにシート分け
  ・シートの縦方向に順次作成
  ・薬袋1枚につき
      - 24行（予定）
      - 3列（幅：2,10,2）
  ・余白：全て0

  class MedBagInfo
    薬袋情報設定クラス
    ファイル名：薬袋_yyyymmdd.xlsx
      yyyymmdd = 処方日

"""
import openpyxl as xl
from openpyxl.styles import Font, Alignment, Border, Side

# 薬袋ファイル設定
FILE_NAME = "薬袋‗"
FILE_EXT = ".xlsx"
# PAPER_SIZE = "A5"
# MARGIN = 0

# 見出し種別
INDEX_ORAL = "[内服薬]"
INDEX_ORAL_ODO = "[内服薬(一包化)]"
INDEX_EXT = "[外用薬]"
INDEX_INJ = "[注射薬]"
INDEX_ASND = "[屯服薬]"
tag_list = [INDEX_ORAL, INDEX_ORAL_ODO, INDEX_EXT,
            INDEX_INJ, INDEX_ASND]

# 薬袋種別
BAG_ORAL = "内　服　薬"
BAG_EXT = "外　用　薬"
BAG_INJ = "注　射　薬"
BAG_ASN = "屯　服　薬"

# 薬袋色設定
COLOR_ORAL = '0070C0'  # 内服
COLOR_EXT = 'FF0000'  # 外用
COLOR_ASN = '00B050'  # 頓服
COLOR_ASA = 'FF0000'  # 朝食後
COLOR_HIRU = '00B050'  # 昼食後
COLOR_YU = '0070C0'  # 夕食後
COLOR_WAKE = 'FC6E04'  # 起床時
COLOR_SLP = '000000'  # 寝る前

# 用法
ODO = "１回　１包"

# 薬局情報/スタンプ欄
PH_NAME = "イルカ薬局　錦町店"
PH_ADDRESS = "小樽市錦町５番１０号"
PH_TEL = "TEL　0134-21-5775"
PH_FAX = "FAX　0134-21-5776"
STAMP_INDEX = "薬剤師"

PT_TITLE = "　様"

# 行設定：基本0開始
MAX_ROW = 24  # 1薬袋の行数
ROW_DATE = 0  # 日付行
ROW_INDEX = 1  # 内服・外用・頓服表示
ROW_PT = 3  # 患者名
ROW_DOSAGE = 5  # 用法用量開始位置（⁺2行）
ROW_MED = 9  # 薬品名開始位置（⁺9行）
ROW_PH = 20  # 薬局名挿入位置
ROW_STAMP = 20  # 印鑑枠INDEX挿入位置
center_list = [ROW_INDEX, ROW_PT, ROW_DOSAGE + 1]
# 行高
HEIGHT_INDEX = 2  # 内服・外用・頓服表示
HEIGHT_PT = 1.7  # 患者名、用法メイン
HEIGHT_DOSAGE = 1  # 用法用量
HEIGHT_DATE = 0.75  # 日付、薬局名
HEIGHT_NORMAL = 0.6  # normal

# 列設定：基本0開始
MAX_COL = 3  # 1薬袋の列数
COL_MAIN = 1  # メイン表記位置
COL_STAMP = 2  # 印鑑枠の列
# 列幅
WIDTH_MAIN = 10
WIDTH_ETC = 2


class MedBag:
    def __init__(self, path, df_list):
        # 初期化、引数処理
        self.dir_path = path
        self.rx_list = df_list
        # 保存ファイル設定
        self.file_name = FILE_NAME
        self.file_path = self.dir_path
        self.wb = xl.Workbook()  # wb新規作成 -> 上書き
        # 処方情報関連
        self.rx_date = ""  # 処方日（YYYY/mm/dd）
        self.pt_num = len(self.rx_list)
        # フォント設定
        self.set_font = Font(name='メイリオ', size=11)
        self.set_center = Alignment(horizontal='center', vertical='center')
        self.set_right = Alignment(horizontal='right', vertical='center')
        # 保存ファイル名設定
        self._set_filename()

    # ===============================================
    # ファイル名設定 & 日付形式変換
    #   ・ファイル名："薬袋_yyyymmdd.xlsx"
    #   ・日付変換：_str_to_date(str)
    # ===============================================
    def _set_filename(self):
        # 日付取得 -> ファイル名設定
        rx = self.rx_list[0]
        r_date = rx.iat[1, 1]
        self.file_name += r_date + FILE_EXT
        self.file_path += "\\" + self.file_name

        # 日付変換
        self._str_to_date(r_date)

    # ===============================================
    # 日付形式変換
    #   yyyymmdd -> yyyy/mm/dd
    # ===============================================
    def _str_to_date(self, d_str):
        from datetime import datetime as dt
        adt = dt.strptime(d_str, '%Y%m%d')
        self.rx_date = adt.strftime('%Y/%m/%d')

    # ===============================================
    # TODO 処方内容取得
    # ===============================================
    def _get_rx_data(self):
        for idx, rx in enumerate(self.rx_list):
            hp = rx.iat[1, 4]  # 病院名
            pt = rx.iat[0, 4]  # 患者名
            # シート作成
            self._create_sheet(pt, idx)

            # TODO 処方情報のまとめこみ処理
            #   【進行中】：データ処理方法かんがえちゅう
            for row in rx:
                # タグをトリガーに処方取得
                # row[] = tag:3, med/dosage:3, num:4, unit:5
                rx_flg = False  # 処方情報開始フラグ
                if row[3] in tag_list:
                    pass

                elif rx_flg:
                    pass

                else:
                    continue

            rx_flg = False

            # 薬袋項目設定
            bag_date = self.rx_date + "　" + hp

    # ===============================================
    # TODO シート作成：患者名
    # ===============================================
    def _create_sheet(self, pt, idx):
        # 患者名シートの作成
        self.wb.create_sheet(title=pt, index=idx)

    # ===============================================
    # TODO 薬袋作成
    # ===============================================
    def make_medbag(self):
        pass

    # ===============================================
    # TODO シート設定
    #   -> 体裁、印刷設定：未テスト
    # ===============================================
    def _sheet_setting(self):

        # シート内設定
        for ws in self.wb.worksheets:
            # フォント名一括設定
            for row in ws:
                for cell in row:
                    cell.font = self.set_font
            # 列幅設定
            # TODO 設定単位確認（Excel既定のcmで設定中）
            ws.column_dimensions['A'].width = WIDTH_ETC
            ws.column_dimensions['B'].width = WIDTH_MAIN
            ws.column_dimensions['C'].width = WIDTH_ETC

            col_main = COL_MAIN + 1  # cell設定が1開始
            col_stamp = COL_STAMP + 1
            # 薬袋1つごと（24行）にループ
            for row in range(0, ws.max_row, 24):
                # 行高設定
                # TODO 設定単位確認（Excel既定のcmで設定中）
                ws.row_dimensions[row].height = HEIGHT_DATE  # 1: date
                ws.row_dimensions[row + ROW_INDEX].height = HEIGHT_INDEX  # 2: index
                ws.row_dimensions[row + ROW_PT].height = HEIGHT_PT  # 4: pt name
                ws.row_dimensions[row + ROW_DOSAGE].height = HEIGHT_DOSAGE  # 6: dosage
                ws.row_dimensions[row + ROW_DOSAGE + 1].height = HEIGHT_PT  # 7: dosage main
                ws.row_dimensions[row + ROW_DOSAGE + 2].height = HEIGHT_DOSAGE  # 8: dosage
                ws.row_dimensions[row + ROW_PH].height = HEIGHT_DATE  # 21: ph name
                # TODO フォント設定
                #   -> 色設定；分岐必要
                # フォント名は別に一括設定
                r = row + 1  # cell設定が1開始
                ws.cell(r, col_main).font = Font(size=14)
                ws.cell(r + ROW_INDEX, col_main).font = Font(color=COLOR_EXT, size=36, bold=True)
                ws.cell(r + ROW_PT, col_main).font = Font(size=28)
                ws.cell(r + ROW_DOSAGE, col_main).font = Font(size=18)
                ws.cell(r + ROW_DOSAGE + 1, col_main).font = Font(color=COLOR_EXT, size=28, bold=True)
                ws.cell(r + ROW_DOSAGE + 2, col_main).font = Font(size=18)
                ws.cell(r + ROW_PH, col_main).font = Font(size=14, bold=True)
                ws.cell(r + ROW_PH, col_stamp).font = Font(size=9)
                # 中央揃え
                for i in center_list:
                    ws.cell(r + i, col_main).alignment = self.set_center
                ws.cell(r + ROW_STAMP, col_stamp).alignment = self.set_center
                # 右寄せ
                ws.cell(r, col_main).alignment = self.set_right
                ws.cell(r + ROW_DOSAGE + 2, col_main).alignment = self.set_right
                # 罫線設定 : stamp
                self._set_border(ws, row)

            # 印刷設定
            self._print_setting(ws)

        self.wb.save(self.file_path)

    # ===============================================
    # 罫線設定
    # ===============================================
    def _set_border(self, ws, r):
        # set side style
        side_def = Side(style='thin', color='000000')

        # set border
        border_top = Border(top=side_def, bottom=side_def, left=side_def, right=side_def)
        border_side = Border(left=side_def, right=side_def)
        border_btm = Border(bottom=side_def, left=side_def, right=side_def)

        # write in sheet
        row = r + ROW_STAMP + 1
        col = COL_STAMP + 1
        ws.cell(row, col).border = border_top
        ws.cell(row + 1, col).border = border_side
        ws.cell(row + 2, col).border = border_side
        ws.cell(row + 3, col).border = border_btm

    # ===============================================
    # 印刷設定
    # ===============================================
    def _print_setting(self, ws):
        wps = ws.page_setup
        # 印刷の向きを設定：縦
        wps.orientation = ws.ORIENTATION_LANDSCAPE
        # 横を１ページ
        wps.fitToWidth = 1
        # 縦を自動
        wps.fitToHeight = 0
        # fitTo属性を有効にする
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        # 用紙サイズを設定
        wps.paperSize = ws.PAPERSIZE_A5

        # 余白
        ws.page_margins.left = 0
        ws.page_margins.right = 0
        ws.page_margins.top = 0
        ws.page_margins.bottom = 0
        ws.page_margins.header = 0
        ws.page_margins.footer = 0
